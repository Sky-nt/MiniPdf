"""Check chart anchor rows for classic113."""
from openpyxl import load_workbook
wb = load_workbook('../MiniPdf.Scripts/output/classic113_chart_sheet.xlsx')
ws = wb.active
print(f'Dimensions: {ws.dimensions}')
print(f'Rows:')
for row in ws.iter_rows(min_row=1, max_row=10, values_only=False):
    vals = [(c.coordinate, c.value) for c in row if c.value is not None]
    if vals:
        print(f'  {vals}')

# Check chart anchors
for chart in ws._charts:
    print(f'\nChart: {chart.title}')
    print(f'  type: {chart.type}')
    if hasattr(chart, 'anchor'):
        print(f'  anchor: {chart.anchor}')
